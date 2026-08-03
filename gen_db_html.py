# -*- coding: utf-8 -*-
# yosou_db_11stores.xlsx -> site/db.html 変換(毎晩の定期タスクから実行)
from openpyxl import load_workbook
import html as H

wb = load_workbook("/root/palazzo-misato/yosou_db_11stores.xlsx", data_only=True)
sheets = ["総合サマリ", "イベント法則_全店", "公約まとめ", "並びデータ", "告知アカウント_全店", "実績ログ_全店"]
tab_ids = ["sum", "law", "koyaku", "narabi", "sns", "log"]
tab_names = ["総合サマリ", "イベント法則", "公約", "並び", "告知網", "実績ログ"]

def sheet_to_table(ws):
    rows = [[("" if c.value is None else str(c.value)) for c in r] for r in ws.iter_rows()]
    rows = [r for r in rows if any(x.strip() for x in r)]
    if not rows: return "<p>データなし</p>"
    ncol = max(len(r) for r in rows)
    while ncol > 0 and all((len(r) <= ncol-1 or not r[ncol-1].strip()) for r in rows): ncol -= 1
    title = rows[0][0] if rows else ""
    hidx = 1
    for i, r in enumerate(rows[1:], 1):
        if sum(1 for x in r[:ncol] if x.strip()) >= 3: hidx = i; break
    out = [f'<div class="ttl">{H.escape(title)}</div>', '<div class="twrap"><table>']
    out.append("<thead><tr>" + "".join(f"<th>{H.escape(x)}</th>" for x in rows[hidx][:ncol]) + "</tr></thead><tbody>")
    for r in rows[hidx+1:]:
        if not any(x.strip() for x in r[:ncol]): continue
        out.append("<tr>" + "".join(f"<td>{H.escape(x)}</td>" for x in r[:ncol]) + "</tr>")
    out.append("</tbody></table></div>")
    return "\n".join(out)

panels = ""; tabs = ""
for i, (sn, tid, tn) in enumerate(zip(sheets, tab_ids, tab_names)):
    ws = wb[sn]
    panels += f'<section id="{tid}" class="panel{" on" if i==0 else ""}">{sheet_to_table(ws)}</section>\n'
    tabs += f'<button class="tab{" on" if i==0 else ""}" data-t="{tid}">{tn}</button>'

page = f"""<!DOCTYPE html>
<html lang="ja"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="manifest.json"><link rel="apple-touch-icon" href="icon-192.png">
<meta name="apple-mobile-web-app-capable" content="yes"><meta name="theme-color" content="#12151c">
<title>法則データベース</title>
<style>
:root{{--bg:#f4f5f7;--card:#fff;--ink:#1a2333;--ink2:#5a6577;--line:#e3e6ec;--accent:#c8102e;--chip:#eef1f6;}}
@media (prefers-color-scheme:dark){{:root{{--bg:#12151c;--card:#1c202a;--ink:#eef1f6;--ink2:#a8b0bf;--line:#2a2f3c;--chip:#262c3a;--accent:#ff5a6e;}}}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:"Hiragino Sans","Yu Gothic",-apple-system,sans-serif;background:var(--bg);color:var(--ink);padding:12px;line-height:1.5}}
.nav{{display:flex;gap:8px;margin-bottom:12px}}
.nav a{{font-size:13px;font-weight:800;text-decoration:none;color:var(--ink2);background:var(--card);border:1px solid var(--line);border-radius:20px;padding:6px 14px}}
.nav a.here{{color:var(--accent);border-color:var(--accent)}}
h1{{font-size:18px;margin-bottom:10px}}
.tabs{{display:flex;gap:6px;overflow-x:auto;padding-bottom:8px;-webkit-overflow-scrolling:touch}}
.tab{{flex:none;font-size:12px;font-weight:700;border:1px solid var(--line);background:var(--card);color:var(--ink2);border-radius:20px;padding:6px 13px;cursor:pointer}}
.tab.on{{background:var(--accent);border-color:var(--accent);color:#fff}}
.panel{{display:none}}.panel.on{{display:block}}
.ttl{{font-size:13px;font-weight:800;color:var(--ink2);margin:10px 0 8px}}
.twrap{{overflow-x:auto;background:var(--card);border:1px solid var(--line);border-radius:12px;-webkit-overflow-scrolling:touch}}
table{{border-collapse:collapse;font-size:11.5px;min-width:640px}}
th{{background:var(--chip);color:var(--ink2);text-align:left;padding:7px 9px;position:sticky;top:0;white-space:nowrap}}
td{{padding:7px 9px;border-top:1px solid var(--line);vertical-align:top;max-width:340px}}
tr td:first-child{{font-weight:700;white-space:nowrap}}
footer{{font-size:10px;color:var(--ink2);margin-top:14px;text-align:center}}
</style></head><body>
<div class="nav"><a href="./">★ 明日の狙い目</a><a class="here" href="db.html">📚 法則データベース</a></div>
<h1>法則データベース(11店舗)</h1>
<input id="q" type="search" placeholder="検索: 店舗・イベント・機種 (例: カバネリ / 2のつく日 / パラ三郷)" style="width:100%;padding:9px 12px;border:1px solid var(--line);border-radius:10px;background:var(--card);color:var(--ink);font-size:13px;margin-bottom:8px">
<div class="tabs" style="margin-bottom:4px" id="quick">
<button class="tab" data-q="周年">周年</button><button class="tab" data-q="ゾロ目">ゾロ目</button><button class="tab" data-q="8のつく日">8の日</button><button class="tab" data-q="2のつく日">2の日</button><button class="tab" data-q="全台系">全台系</button><button class="tab" data-q="ジャグ">ジャグラー</button><button class="tab" data-q="カバネリ">カバネリ</button><button class="tab" data-q="喰種">喰種</button><button class="tab" data-q="">クリア</button>
</div>
<div class="tabs">{tabs}</div>
{panels}
<footer>yosou_db_11stores.xlsx より自動生成 ｜ 毎晩の定期更新で最新化</footer>
<script>
document.querySelectorAll('.tab').forEach(b=>b.addEventListener('click',()=>{{
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('on'));
  document.querySelectorAll('.panel').forEach(x=>x.classList.remove('on'));
  b.classList.add('on'); document.getElementById(b.dataset.t).classList.add('on');
}}));
const q=document.getElementById('q');
function filt(){{const v=q.value.trim().toLowerCase();
 document.querySelectorAll('.panel table tbody tr').forEach(tr=>{{ tr.style.display = (!v || tr.textContent.toLowerCase().includes(v)) ? '' : 'none'; }});
 document.querySelectorAll('.panel').forEach(p=>{{const vis=[...p.querySelectorAll('tbody tr')].filter(t=>t.style.display!=='none').length; }});
}}
q.addEventListener('input',filt);
document.querySelectorAll('#quick .tab').forEach(b=>b.addEventListener('click',()=>{{q.value=b.dataset.q;filt();}}));
if('serviceWorker' in navigator){{navigator.serviceWorker.register('sw.js');}}
</script></body></html>"""
with open("/root/palazzo-misato/site/db.html", "w", encoding="utf-8") as f: f.write(page)
print("db.html regenerated")
